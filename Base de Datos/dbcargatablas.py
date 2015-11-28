from dbconnect import connection
import openpyxl

wb = openpyxl.load_workbook("Proveedores.xlsx")
hoja = wb.get_sheet_names()
hojaproveedores = wb.get_sheet_by_name(hoja[0])
print("Archivo Proveedores Abierto: ", wb)
print(hojaproveedores)

c, conn = connection()

for renglon in range(2,100000):
    if hojaproveedores.cell(row=renglon, column=1).value is None:
        break
    Nom = hojaproveedores.cell(row=renglon, column=1).value
    c.execute("SELECT * FROM Proveedores WHERE NOMBRE = (%s)",
              (Nom))
    rows = c.fetchall()
    if len(rows) == 0:
        Num = hojaproveedores.cell(row=renglon, column=2).value
        Dir = hojaproveedores.cell(row=renglon, column=3).value
        Ciu = hojaproveedores.cell(row=renglon, column=4).value
        Tel = hojaproveedores.cell(row=renglon, column=5).value
        RFC = hojaproveedores.cell(row=renglon, column=6).value
        Com = hojaproveedores.cell(row=renglon, column=7).value
        print('Instertando', Nom)
        c.execute("INSERT INTO Proveedores (NOMBRE, NUMERO, DIRECCION, CIUDAD, TELEFONOS, RFC, COMENTARIOS) VALUES (%s, %s, %s, %s, %s, %s, %s)",
                  (Nom, Num, Dir, Ciu, Tel, RFC, Com))
    else:
        print('Ya existe', Nom, rows)
    
        conn.commit()
conn.close()
    
    
    
